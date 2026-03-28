"""
reports_panel.py - Sales reports with CSV/Excel export (Admin only).
"""

import customtkinter as ctk
from tkinter import ttk, filedialog, messagebox
import tkinter as tk
from datetime import datetime
import csv
import database as db
from theme import *
from widgets import (section_label, make_tree, populate_tree,
                     accent_btn, neutral_btn, blue_btn,
                     stat_card, dim_label, show_toast)


class ReportsPanel(ctk.CTkFrame):
    def __init__(self, master, **kw):
        super().__init__(master, fg_color=CLR_BG, corner_radius=0, **kw)
        self._build()
        self.refresh()

    def _build(self):
        # ── Header ───────────────────────────────────────────
        hdr = ctk.CTkFrame(self, fg_color="transparent")
        hdr.pack(fill="x", padx=PAD, pady=(PAD, 0))
        section_label(hdr, "📊  Sales Reports").pack(side="left")

        export_row = ctk.CTkFrame(hdr, fg_color="transparent")
        export_row.pack(side="right")
        blue_btn(export_row,  "⬇ Export CSV",   self._export_csv,   width=130).pack(
            side="left", padx=4)
        accent_btn(export_row,"⬇ Export Excel", self._export_excel, width=140).pack(
            side="left", padx=4)
        neutral_btn(export_row,"↺ Refresh",      self.refresh,       width=100).pack(
            side="left", padx=4)

        # ── KPI cards ────────────────────────────────────────
        kpi = ctk.CTkFrame(self, fg_color="transparent")
        kpi.pack(fill="x", padx=PAD, pady=(PAD, 0))

        self.kpi_frames = {}
        for key, label, color in [
            ("total_sales",    "Total Transactions", CLR_ACCENT2),
            ("total_revenue",  "Total Revenue",      CLR_ACCENT),
            ("today_sales",    "Today's Sales",      CLR_WARNING),
            ("today_revenue",  "Today's Revenue",    "#A78BFA"),
        ]:
            card, lbl = stat_card(kpi, label, "—", color=color, width=190)
            card.pack(side="left", padx=6, pady=6)
            self.kpi_frames[key] = lbl

        # ── Tabs ─────────────────────────────────────────────
        tabs = ctk.CTkTabview(self, fg_color=CLR_SURFACE,
                              segmented_button_fg_color=CLR_SURFACE2,
                              segmented_button_selected_color=CLR_ACCENT,
                              segmented_button_selected_hover_color=CLR_ACCENT_DARK,
                              segmented_button_unselected_color=CLR_SURFACE2,
                              text_color=CLR_TEXT,
                              corner_radius=10)
        tabs.pack(fill="both", expand=True, padx=PAD, pady=PAD)
        tabs.add("All Sales")
        tabs.add("Daily Summary")
        tabs.add("Best Sellers")

        self._build_all_sales(tabs.tab("All Sales"))
        self._build_daily(tabs.tab("Daily Summary"))
        self._build_bestsellers(tabs.tab("Best Sellers"))

    # ── All Sales tab ─────────────────────────────────────────
    def _build_all_sales(self, parent):
        # Date filter row
        filt = ctk.CTkFrame(parent, fg_color="transparent")
        filt.pack(fill="x", padx=8, pady=8)

        ctk.CTkLabel(filt, text="Filter by date:", font=FONT_BODY,
                     text_color=CLR_TEXT_DIM).pack(side="left")
        self.date_var = ctk.StringVar(
            value=datetime.now().strftime("%Y-%m-%d"))
        de = ctk.CTkEntry(filt, textvariable=self.date_var,
                          width=120, font=FONT_BODY,
                          fg_color=CLR_SURFACE2, border_color=CLR_BORDER,
                          text_color=CLR_TEXT, height=32, corner_radius=6)
        de.pack(side="left", padx=8)
        accent_btn(filt, "Filter", self._filter_date, width=80).pack(side="left")
        neutral_btn(filt, "Show All", self._show_all, width=90).pack(
            side="left", padx=6)

        cols = ("id", "date", "total", "cashier")
        heads = ("Sale ID", "Date & Time", "Total (SAR)", "Cashier")
        tf, self.all_tree = make_tree(parent, cols, height=15)
        tf.pack(fill="both", expand=True, padx=8, pady=(0, 8))
        for col, head, w in zip(cols, heads, (70, 200, 110, 130)):
            self.all_tree.heading(col, text=head)
            self.all_tree.column(col, width=w, minwidth=50)

    def _filter_date(self):
        date = self.date_var.get().strip()
        try:
            datetime.strptime(date, "%Y-%m-%d")
        except ValueError:
            messagebox.showwarning("Invalid Date",
                                   "Use format YYYY-MM-DD"); return
        rows = db.get_all_sales(date_filter=date)
        self._populate_sales(rows)

    def _show_all(self):
        rows = db.get_all_sales()
        self._populate_sales(rows)

    def _populate_sales(self, rows):
        self.all_tree.delete(*self.all_tree.get_children())
        for i, r in enumerate(rows):
            tag = "odd" if i % 2 == 0 else "even"
            self.all_tree.insert("", "end", values=(
                f"SALE-{r['id']:05d}", r["date"][:16],
                f"SAR {r['total']:.2f}", r["cashier"] or "—"
            ), tags=(tag,))

    # ── Daily Summary tab ─────────────────────────────────────
    def _build_daily(self, parent):
        cols = ("day", "transactions", "revenue")
        heads = ("Date", "Transactions", "Revenue (SAR)")
        tf, self.daily_tree = make_tree(parent, cols, height=18)
        tf.pack(fill="both", expand=True, padx=8, pady=8)
        for col, head, w in zip(cols, heads, (140, 130, 150)):
            self.daily_tree.heading(col, text=head)
            self.daily_tree.column(col, width=w, minwidth=60)

    # ── Best Sellers tab ──────────────────────────────────────
    def _build_bestsellers(self, parent):
        cols = ("name", "units_sold", "revenue")
        heads = ("Product", "Units Sold", "Revenue (SAR)")
        tf, self.best_tree = make_tree(parent, cols, height=18)
        tf.pack(fill="both", expand=True, padx=8, pady=8)
        for col, head, w in zip(cols, heads, (200, 120, 150)):
            self.best_tree.heading(col, text=head)
            self.best_tree.column(col, width=w, minwidth=60)

    # ── Refresh ───────────────────────────────────────────────
    def refresh(self):
        summary = db.get_summary()
        today   = datetime.now().strftime("%Y-%m-%d")
        t_rows  = db.get_all_sales(date_filter=today)
        t_rev   = sum(r["total"] for r in t_rows)

        self.kpi_frames["total_sales"].configure(
            text=str(summary["cnt"]))
        self.kpi_frames["total_revenue"].configure(
            text=f"SAR {summary['revenue']:.2f}")
        self.kpi_frames["today_sales"].configure(
            text=str(len(t_rows)))
        self.kpi_frames["today_revenue"].configure(
            text=f"SAR {t_rev:.2f}")

        # All sales (default: today)
        self._filter_date()

        # Daily
        self.daily_tree.delete(*self.daily_tree.get_children())
        for i, r in enumerate(db.get_daily_sales()):
            tag = "odd" if i % 2 == 0 else "even"
            self.daily_tree.insert("", "end", values=(
                r["day"], r["transactions"],
                f"SAR {r['revenue']:.2f}"
            ), tags=(tag,))

        # Best sellers
        self.best_tree.delete(*self.best_tree.get_children())
        for i, r in enumerate(db.get_best_sellers()):
            tag = "odd" if i % 2 == 0 else "even"
            self.best_tree.insert("", "end", values=(
                r["name"], r["units_sold"],
                f"SAR {r['revenue']:.2f}"
            ), tags=(tag,))

    # ── Export ────────────────────────────────────────────────
    def _export_csv(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv")],
            initialfile=f"sales_report_{datetime.now():%Y%m%d}.csv"
        )
        if not path:
            return
        rows = db.get_all_sales()
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["Sale ID", "Date", "Total (SAR)", "Cashier"])
            for r in rows:
                w.writerow([f"SALE-{r['id']:05d}", r["date"],
                             f"{r['total']:.2f}", r["cashier"] or ""])
        show_toast(self, f"✓  CSV exported successfully")

    def _export_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            messagebox.showerror("Missing Package",
                                 "Run: pip install openpyxl"); return

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"sales_report_{datetime.now():%Y%m%d}.xlsx"
        )
        if not path:
            return

        wb = openpyxl.Workbook()

        # ── Sheet 1: All Sales ──
        ws1 = wb.active
        ws1.title = "All Sales"
        hdr_fill  = PatternFill("solid", fgColor="22C55E")
        hdr_font  = Font(bold=True, color="FFFFFF")
        headers   = ["Sale ID", "Date", "Total (SAR)", "Cashier"]
        for ci, h in enumerate(headers, 1):
            cell = ws1.cell(1, ci, h)
            cell.fill = hdr_fill
            cell.font = hdr_font
        for r in db.get_all_sales():
            ws1.append([f"SALE-{r['id']:05d}", r["date"],
                        r["total"], r["cashier"] or ""])
        for col in ws1.columns:
            ws1.column_dimensions[col[0].column_letter].width = 20

        # ── Sheet 2: Daily Summary ──
        ws2 = wb.create_sheet("Daily Summary")
        for ci, h in enumerate(["Date", "Transactions", "Revenue (SAR)"], 1):
            cell = ws2.cell(1, ci, h)
            cell.fill = hdr_fill
            cell.font = hdr_font
        for r in db.get_daily_sales():
            ws2.append([r["day"], r["transactions"], r["revenue"]])

        # ── Sheet 3: Best Sellers ──
        ws3 = wb.create_sheet("Best Sellers")
        for ci, h in enumerate(["Product", "Units Sold", "Revenue (SAR)"], 1):
            cell = ws3.cell(1, ci, h)
            cell.fill = hdr_fill
            cell.font = hdr_font
        for r in db.get_best_sellers():
            ws3.append([r["name"], r["units_sold"], r["revenue"]])

        wb.save(path)
        show_toast(self, "✓  Excel report exported successfully")
